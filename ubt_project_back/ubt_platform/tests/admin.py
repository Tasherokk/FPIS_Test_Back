from collections import defaultdict
from datetime import datetime

import nested_admin
from django.contrib import admin, messages
from django.contrib.admin import DateFieldListFilter
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.forms import AdminPasswordChangeForm
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django_summernote.widgets import SummernoteWidget
from django import forms
from .models import Subject, Question, Answer, MatchingPair, TestResult, SubjectResult, CustomUser, School
from django.urls import reverse, path
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.styles import Font


# Кастомная форма для Answer с использованием Summernote
class AnswerForm(forms.ModelForm):
    class Meta:
        model = Answer
        fields = '__all__'
        widgets = {
            'text': SummernoteWidget(),  # Используем Summernote для редактирования текста
        }


class AnswerInline(nested_admin.NestedTabularInline):
    model = Answer
    extra = 0
    max_num = 6
    form = AnswerForm  # Применяем форму с Summernote


class MatchingPairForm(forms.ModelForm):
    class Meta:
        model = MatchingPair
        fields = '__all__'
        widgets = {
            'left_side_1': SummernoteWidget(),
            'left_side_2': SummernoteWidget(),  # Используем Summernote для редактирования текста
            'right_option_1': SummernoteWidget(),
            'right_option_2': SummernoteWidget(),
            'right_option_3': SummernoteWidget(),
            'right_option_4': SummernoteWidget(),
        }


class MatchingPairInline(nested_admin.NestedStackedInline):
    model = MatchingPair
    extra = 0
    max_num = 1
    form = MatchingPairForm  # Применяем форму с Summernote


# Кастомная форма для Question с использованием Summernote
class QuestionForm(forms.ModelForm):
    class Meta:
        model = Question
        fields = '__all__'
        widgets = {
            'text': SummernoteWidget(),  # Используем Summernote для редактирования текста
        }


class QuestionInline(nested_admin.NestedStackedInline):
    model = Question
    extra = 0
    max_num = 40
    inlines = [AnswerInline, MatchingPairInline]
    form = QuestionForm  # Применяем форму с Summernote


class SubjectAdmin(nested_admin.NestedModelAdmin):
    inlines = [QuestionInline]
    list_display = ['name', 'variant', 'is_active']
    list_filter = ['name', 'variant', 'is_active']
    list_editable = ['is_active']


# Админ-класс для Question с поддержкой Summernote
class QuestionAdmin(nested_admin.NestedModelAdmin):
    inlines = [AnswerInline, MatchingPairInline]
    form = QuestionForm


class MatchingPairAdmin(nested_admin.NestedModelAdmin):
    form = MatchingPairForm


class SubjectResultInline(nested_admin.NestedStackedInline):
    model = SubjectResult
    exclude = ['subject', 'score']


class TotalScoreFilter(admin.SimpleListFilter):
    title = _('total score')  # Name for the filter in the admin interface
    parameter_name = 'total_score'  # URL parameter for the filter

    def lookups(self, request, model_admin):
        return (
            ('increasing', _('Возрастающий')),
            ('decreasing', _('Убывающий')),
        )

    def queryset(self, request, queryset):
        if self.value() == 'increasing':
            return queryset.order_by('total_score')
        if self.value() == 'decreasing':
            return queryset.order_by('-total_score')
        return queryset


class TestResultAdmin(nested_admin.NestedModelAdmin):
    change_list_template = "admin/change_list.html"

    date_hierarchy = 'date_taken'

    list_display = ['user_full_name', 'user_iin', 'user_school', 'total_score', 'date_taken']
    list_filter = [
        ('date_taken', DateFieldListFilter),
        'user__usage_type',
        ('user__created_at', DateFieldListFilter),
        'user__school',
        'user',
        TotalScoreFilter
    ]
    ordering = ['-date_taken', 'user__full_name']
    inlines = [SubjectResultInline]
    readonly_fields = ['id', 'total_score', 'date_taken', 'user']

    def user_full_name(self, obj):
        return obj.user.full_name

    user_full_name.short_description = 'ФИО'

    def user_iin(self, obj):
        return obj.user.iin

    user_iin.short_description = 'ИИН'

    def user_school(self, obj):
        return obj.user.school.name if obj.user.school else "—"

    user_school.short_description = 'Школа'

    SUBJECT_CODE_TO_RU = {
        'HIS': 'История',
        'RL': 'Грам. чтения',
        'ML': 'Матем. грам.',
        'BIO': 'Биология',
        'CHE': 'Химия',
        'MAT': 'Математика',
        'PHY': 'Физика',
        'GEO': 'География',
        'LF': 'Право',
        'FL': 'Иностр. язык',
        'KZ': 'Казахский яз.',
        'KL': 'Казахская лит.',
        'INF': 'Информатика',
        'WHI': 'Всемир. история',
        'RU': 'Русский язык',
        'RUL': 'Русская лит.',
    }

    DEFAULT_CODES = ['HIS', 'RL', 'ML']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("export-excel/", self.admin_site.admin_view(self.export_excel), name="export_excel"),
        ]
        return custom_urls + urls

    def export_excel(self, request):
        if request.method == "POST":
            school_id = request.POST.get("school")
            date_str = request.POST.get("date")

            if not school_id or not date_str:
                return render(request, "admin/export_results.html", {
                    "schools": School.objects.all(),
                    "error": "Выберите школу и дату!"
                })

            try:
                selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return render(request, "admin/export_results.html", {
                    "schools": School.objects.all(),
                    "error": "Некорректный формат даты!"
                })

            school = School.objects.filter(id=school_id).first()
            if not school:
                return render(request, "admin/export_results.html", {
                    "schools": School.objects.all(),
                    "error": "Выбранная школа не найдена!"
                })

            # Берём результаты тестов за выбранную дату по заданной школе
            test_results = (
                TestResult.objects.filter(
                    user__school_id=school_id,
                    date_taken__date=selected_date
                )
                .select_related("user")
                .order_by("user", "-total_score")
                .distinct("user")
            )

            if not test_results.exists():
                return render(request, "admin/export_results.html", {
                    "schools": School.objects.all(),
                    "error": "Нет данных для выбранной школы и даты!"
                })

            # -----------------------
            # Готовим группировку по "паре дополнительных предметов"
            # -----------------------
            # Для каждого test_result определяем, какие предметы у него были (кроме 3 статичных),
            # и группируем результаты по этим парам.
            # Например, (BIO, CHE) -> [список TestResult], (INF, MAT) -> [список TestResult], ...
            group_to_results = defaultdict(list)

            for tr in test_results:
                # Получаем все SubjectResult для данного TestResult
                subject_res_qs = tr.subject_results.select_related('subject')

                # Выделим коды предметов
                codes = [sr.subject.name for sr in subject_res_qs]
                # Оставляем «дополнительные» предметы = те, что не в DEFAULT_CODES
                optional_codes = [c for c in codes if c not in self.DEFAULT_CODES]

                # На случай, если вдруг там нет 2х предметов, подстрахуемся
                # Сортируем, чтобы пара (BIO, CHE) была тем же, что и (CHE, BIO)
                optional_codes = sorted(optional_codes)

                # Превращаем в tuple, чтобы можно было использовать как ключ в dict
                pair_key = tuple(optional_codes)
                group_to_results[pair_key].append(tr)

            # Теперь у нас есть словарь: {('BIO','CHE'): [...], ('INF','MAT'): [...], ...}

            # Создаем Excel
            wb = Workbook()
            # Удалим стартовый лист "Sheet", чтобы не мешался,
            # или можем просто переименовать, но удобнее удалить и создавать с нуля
            default_sheet = wb.active
            wb.remove(default_sheet)

            # Для каждого "pair_key" (уникальная пара доп. предметов) делаем свой лист
            for pair_key, results_in_group in group_to_results.items():
                # Если по каким-то причинам пара пустая или одна, формируем название как есть
                if not pair_key:
                    sheet_name = "Без доп. предметов"
                else:
                    # Пример: "BIO+CHE"
                    sheet_name = "+".join(pair_key)
                ws = wb.create_sheet(title=sheet_name[:31])  # Ограничение Excel в 31 символ

                # Заголовок:
                #  1) Школа
                #  2) Имя пользователя
                #  3) Общий балл
                #  4) Дата прохождения
                #  5) 3 столбца для фиксированных предметов
                #  6) столбцы для предметов из pair_key
                headers = [
                    "Название школы",
                    "Имя пользователя",
                    "Общий балл",
                    "Дата прохождения",
                    # Три статичных предмета
                    self.SUBJECT_CODE_TO_RU.get('HIS', 'HIS'),
                    self.SUBJECT_CODE_TO_RU.get('RL', 'RL'),
                    self.SUBJECT_CODE_TO_RU.get('ML', 'ML')
                ]

                # Добавляем «дополнительные» предметы в заголовок
                for code in pair_key:
                    headers.append(self.SUBJECT_CODE_TO_RU.get(code, code))

                ws.append(headers)

                # Делаем заголовки жирными
                bold_font = Font(bold=True)
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num).font = bold_font

                # Заполняем строки данными
                for test in results_in_group:
                    row = []
                    # 1) Название школы
                    row.append(test.user.school.name if test.user.school else "—")
                    # 2) Имя пользователя
                    row.append(test.user.full_name)
                    # 3) Общий балл
                    row.append(test.total_score)
                    # 4) Дата прохождения
                    row.append(test.date_taken.strftime("%Y-%m-%d %H:%M:%S"))

                    # Нужно вытащить баллы по каждому предмету
                    # Превратим список SubjectResult в словарь {код:балл}
                    sr_qs = test.subject_results.select_related('subject')
                    code_to_score = {}
                    for sr in sr_qs:
                        code_to_score[sr.subject.name] = sr.score

                    # Добавляем 3 фиксированных предмета
                    for def_code in self.DEFAULT_CODES:
                        row.append(code_to_score.get(def_code, 0))

                    # Добавляем баллы по выбранным предметам
                    for opt_code in pair_key:
                        row.append(code_to_score.get(opt_code, 0))

                    ws.append(row)

                    current_row = ws.max_row
                    ws.cell(row=current_row, column=2).font = bold_font

                    # Автоматическая настройка ширины столбцов
                    for column_cells in ws.columns:
                        max_length = 0
                        column = column_cells[0].column_letter
                        for cell in column_cells:
                            try:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                            except:
                                pass
                        ws.column_dimensions[column].width = max_length + 2

            # Формируем имя файла, подставляя школу и дату
            school_name_cleaned = "".join(c if c.isalnum() else "_" for c in school.name)
            file_name = f"test_results_{school_name_cleaned}_{selected_date}.xlsx"

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            wb.save(response)
            return response

        # Если не POST, рендерим форму
        return render(request, "admin/export_results.html", {"schools": School.objects.all()})

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["export_excel_url"] = "export-excel/"
        return super().changelist_view(request, extra_context)


class CustomUserCreationForm(forms.ModelForm):
    class Meta:
        model = CustomUser
        fields = ('full_name', 'iin', 'usage_type')

    def save(self, commit=True):
        user = super().save(commit=False)
        user.set_password(self.cleaned_data["iin"])
        if commit:
            user.save()
        return user


class CustomUserChangeForm(forms.ModelForm):
    class Meta:
        model = CustomUser
        fields = ('full_name', 'is_active', 'created_at', 'usage_type')


class CustomUserAdmin(UserAdmin):
    form = CustomUserChangeForm
    add_form = CustomUserCreationForm
    model = CustomUser
    change_password_form = AdminPasswordChangeForm

    date_hierarchy = 'created_at'

    list_display = ('iin', 'full_name', 'school', 'usage_type', 'created_at', 'is_active')
    list_filter = ('iin', 'full_name', 'is_active', 'usage_type', ('created_at', DateFieldListFilter), 'school')
    search_fields = ('iin', 'full_name')
    ordering = ('full_name',)

    fieldsets = (
        (None, {'fields': ('iin', 'password', 'password_link')}),
        ('Personal info', {'fields': ('full_name', 'school', 'created_at', 'usage_type')}),
        ('Permissions', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions',)}),
    )

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('iin', 'full_name', 'school', 'usage_type', 'is_active', 'is_staff', 'is_superuser', 'groups',
                       'user_permissions',),
        }),
    )

    filter_horizontal = ('groups', 'user_permissions')

    readonly_fields = ['password', 'password_link']

    actions = ['activate_selected_users', 'deactivate_selected_users']

    def activate_selected_users(self, request, queryset):
        updated = queryset.update(is_active=True, created_at=timezone.now())
        self.message_user(request, f"Активировано {updated} пользователь(ей).", level=messages.SUCCESS)

    activate_selected_users.short_description = "Активировать выбранных пользователей"

    def deactivate_selected_users(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f"Деактивировано {updated} пользователь(ей).", level=messages.SUCCESS)

    deactivate_selected_users.short_description = "Деактивировать выбранных пользователей"

    def password_link(self, obj):
        """Добавляем ссылку для смены пароля."""
        if obj.id:
            url = reverse('admin:auth_user_password_change', args=[obj.id])
            return format_html('<a href="{}">{}</a>', url, _('Сменить пароль'))
        return _("Пароль ещё не задан")

    password_link.short_description = "Изменить пароль"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('activate_users/', self.admin_site.admin_view(self.activate_users), name='activate_users'),
            path('deactivate_users/', self.admin_site.admin_view(self.deactivate_users), name='deactivate_users'),
            path('delete_users/', self.admin_site.admin_view(self.delete_users), name='delete_users'),
        ]
        return custom_urls + urls

    def activate_users(self, request):
        return self.process_users(request, action='activate')

    def deactivate_users(self, request):
        return self.process_users(request, action='deactivate')

    def delete_users(self, request):
        return self.process_users(request, action='delete')

    def process_users(self, request, action='activate'):
        if request.method == "POST":
            iin_list = request.POST.get('iin_list')
            if iin_list:
                iins = [iin.strip() for iin in iin_list.splitlines() if iin.strip()]
                users = CustomUser.objects.filter(iin__in=iins).exclude(is_staff=True).exclude(is_superuser=True)
                if users.exists():
                    if action == 'activate':
                        users.update(is_active=True, created_at=timezone.now())
                        messages.success(request, "Пользователи с указанными ИИН были Активированы.")
                    elif action == 'deactivate':
                        users.update(is_active=False)
                        messages.success(request, "Пользователи с указанными ИИН были Деактивированы.")
                    elif action == 'delete':
                        count = users.count()
                        users.delete()
                        messages.success(request, f"Удалено {count} пользователей.")
                else:
                    messages.warning(request, "Ни один из указанных ИИН не найден.")
            else:
                messages.error(request, "Введите хотя бы один ИИН.")
            return redirect("..")

        action_name = {
            'activate': _("Активировать пользователей"),
            'deactivate': _("Деактивировать пользователей"),
            'delete': _("Удалить пользователей")
        }.get(action, _("Изменение статуса пользователей"))

        context = {
            'title': action_name,
            'activate': (action == 'activate'),
            'deactivate': (action == 'deactivate'),
            'delete': (action == 'delete'),
            'opts': self.model._meta,
        }
        return render(request, 'admin/custom_user_admin.html', context)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['custom_buttons'] = format_html(
            '<div style="margin-bottom: 10px;">'
            '<a class="button" style="margin-right: 10px;" href="{}">{}</a>'
            '<a class="button" style="margin-right: 10px;" href="{}">{}</a>'
            '<a class="button" href="{}">{}</a>'
            '</div>',
            reverse('admin:activate_users'), _('Активировать пользователей'),
            reverse('admin:deactivate_users'), _('Деактивировать пользователей'),
            reverse('admin:delete_users'), _('Удалить пользователей')
        )
        return super().changelist_view(request, extra_context=extra_context)


admin.site.register(Subject, SubjectAdmin)
admin.site.register(Question, QuestionAdmin)
admin.site.register(Answer)
admin.site.register(MatchingPair, MatchingPairAdmin)
admin.site.register(TestResult, TestResultAdmin)
admin.site.register(CustomUser, CustomUserAdmin)
admin.site.register(School)
